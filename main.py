import os
import pandas as pd
import xlsxwriter
import copy
import openpyxl
from art import text2art


def choose_file():
    file_name = input('Напишите название файла вместе с расширением(enter - input.xlsx): ')

    if not file_name:
        file_name = 'input.xlsx'

    if file_name.endswith('.xlsx') and os.path.exists(file_name):
        excel_file = openpyxl.load_workbook(file_name)
        excel_sheet = excel_file.active

        table = []
        for i in range(excel_sheet.max_row - 1):
            row = []
            for col in excel_sheet.iter_cols(1, excel_sheet.max_column - 1):
                row.append(col[i].value)
            table.append(row)

        return False, table
    else:
        print('Неправильный формат файла.')
        return True


def console_table():
    print('Пишите значения не разделяя их знаками, а вместо пустых значений пишите 0')
    print()

    table = []

    for raw in range(1, 10):
        print(f'Введите {raw} строку в таблице:')
        items = input()

        if len(items) == 9 and items.isdigit():
            to_tabel = list(items)
            to_tabel = map(lambda x: int(x), to_tabel)

            print()

            table.append(to_tabel)
        else:
            print('Неверное значение!\n')
            return True, []

    return False, table


def zeros_and_empty_to_list_digits(table):
    """
    Заменяет все ноли в таблице списком с цифрами от 1 до 9.

    :param table:
    :return:
    """
    new_table = []
    for line in table:
        line = [[i for i in range(1, 10)] if not x else x for x in line]
        new_table.append(line)

    return new_table


def get_cols(table):
    """
    Возвращает список из столбцов таблицы

    :param table:
    :return:
    """
    cols = []
    for i in range(9):
        col = []
        for j in range(9):
            col.append(table[j][i])
        cols.append(col)

    return cols


def get_squads(table):
    """
    Создает список из элементов таблицы, разделенные на квадраты 3х3, как в судоку.
    Возвращает таблицу элементов и таблицу с соответствующими им индексами во входной таблице.

    :param table:
    :return:
    """
    squads = []
    squad_indexes = []
    for i in range(0, 7, 3):
        for j in range(0, 7, 3):
            squad = []
            squad_index = []
            for k in range(3):
                for m in range(3):
                    squad.append(table[i + k][j + m])
                    squad_index.append([i + k, j + m])
            squads.append(squad)
            squad_indexes.append(squad_index)

    return squads, squad_indexes


def item_in_list(items, key):
    """
    Принимается группа элементов и определенное значение.
    Если в группе есть элемент list с данным числом, то число удаляется.

    :param items:
    :param key:
    :return:
    """
    for i in range(9):
        if isinstance(items[i], list):
            if key in items[i]:
                items[i].remove(key)


def open_list_with_one_item(items):
    """
    Ищет внутри группы элементов списки с одним значением и ставит вместо списка данное значение.

    :param items:
    :return:
    """
    for i in range(9):
        if isinstance(items[i], list):
            if len(items[i]) == 1:
                items[i] = items[i][0]


def only_one_option_in_line(items, item):
    """
    Принимает группу элементов и искомый элемент.
    Если искомый элемент единственный в группе, то возвращает его индекс
    :param items:
    :param item:
    :return:
    """
    count = 0
    indx = None
    for i in range(9):
        if isinstance(items[i], list):
            if item in items[i]:
                count += 1
                indx = i

    if count == 1:
        return indx


def two_pairs(items, item):
    """
    Получает группу элементов и элемент содержащий список из двух цифр.
    Удаляет из всех других списков полученные две цифры, оставляя только те элементы,
    который полностью совпадают с заданными.
    :param items:
    :param item:
    :return:
    """
    for i in range(9):
        if isinstance(items[i], list):
            if items[i] != item:
                if item[0] in items[i]:
                    items[i].remove(item[0])
                if item[1] in items[i]:
                    items[i].remove(item[1])


def two_pairs_line(items, i):
    """
    Получает таблицу элементов и порядковый номер строки. Находит два одиннаковых спаренных значения.
    :param items:
    :param i:
    :return:
    """
    save_items = []
    for j in range(9):
        if isinstance(items[i][j], list):
            if len(items[i][j]) == 2:
                if items[i][j] in save_items:
                    two_pairs(items[i], items[i][j])
                    save_items.remove(items[i][j])
                else:
                    save_items.append(items[i][j])


def main():
    """
    Стартовое меню
    """
    table = []
    choose_mode = True
    while choose_mode:
        mode = input('Для решения судоку выберите способ ввода данных: [1] - Указать xlsx файл, '
                     '[2] - Ввести через коммандную строку, '
                     '[3] - Выйти: ')
        print()
        if mode == '1':
            choose_mode, table = choose_file()
        elif mode == '2':
            choose_mode, table = console_table()
        elif mode == '3':
            return

    table = zeros_and_empty_to_list_digits(table)
    count = 0
    old_table = []

    while old_table != table:
        count += 1
        old_table = copy.deepcopy(table)
        cols = get_cols(table)
        squads, squad_index = get_squads(table)

        """
        Удаление записей о числах
        """
        for digit in range(1, 10):
            for i in range(9):
                if digit in table[i]:
                    item_in_list(table[i], digit)
                if digit in cols[i]:
                    item_in_list(cols[i], digit)
                if digit in squads[i]:
                    item_in_list(squads[i], digit)

        for i in range(9):
            open_list_with_one_item(table[i])
            open_list_with_one_item(cols[i])
            open_list_with_one_item(squads[i])

        for digit in range(1, 10):
            for i in range(9):
                if digit not in table[i]:
                    indx = only_one_option_in_line(table[i], digit)
                    if indx:
                        table[i][indx] = digit

            for i in range(9):
                if digit not in cols[i]:
                    indx = only_one_option_in_line(cols[i], digit)
                    if indx:
                        table[indx][i] = digit
    
            for i in range(9):
                if digit not in squads[i]:
                    indx = only_one_option_in_line(squads[i], digit)
                    if indx:
                        x, y = squad_index[i][indx]
                        table[x][y] = digit
        
        for i in range(9):
            two_pairs_line(table, i)
            two_pairs_line(cols, i)
            two_pairs_line(squads, i)

    print(f'Количество циклов: {count}')
    excel = pd.DataFrame(table)
    print(excel.to_string())
    print()

    print('Результат сохранен в файле: output.xlsx')
    writer = pd.ExcelWriter('output.xlsx', engine='xlsxwriter')
    excel.to_excel(writer, 'Sheet1')
    writer._save()


if __name__ == '__main__':
    text_art = text2art('SUDOKU SOLUTION')
    print(text_art)

    main()
